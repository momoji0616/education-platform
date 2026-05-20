import os
import re
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, TimeoutError
from datetime import datetime
from typing import Any

import dashscope
import pandas as pd
from openpyxl import load_workbook
from qdrant_client import QdrantClient
from qdrant_client.http import models as qmodels
from sklearn.feature_extraction.text import HashingVectorizer

from database import Base, SessionLocal, engine
from llm_client import call_deepseek_chat, is_deepseek_available
from models import RagDataset

# 创建数据库表（如果不存在）
Base.metadata.create_all(bind=engine)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
QDRANT_PATH = os.path.join(BASE_DIR, "qdrant_db")
QDRANT_COLLECTION = "excel_data"
VECTOR_SIZE = 384
ACTIVE_FLAG = "0"
DELETED_FLAG = "1"
SUPPORTED_EXCEL_EXT = {".xlsx", ".xls"}
SUPPORTED_TXT_EXT = ".txt"
DEEPSEEK_TEXT_MODEL = "deepseek-v4-flash"

qdrant_client = QdrantClient(path=QDRANT_PATH)
vectorizer = HashingVectorizer(
    n_features=VECTOR_SIZE,
    alternate_sign=False,
    norm="l2",
    analyzer="char",
    ngram_range=(2, 4),
)


def _ensure_qdrant_collection():
    try:
        qdrant_client.get_collection(QDRANT_COLLECTION)
    except Exception:
        qdrant_client.create_collection(
            collection_name=QDRANT_COLLECTION,
            vectors_config=qmodels.VectorParams(size=VECTOR_SIZE, distance=qmodels.Distance.COSINE),
        )


def _embed_texts(texts):
    if not texts:
        return []
    matrix = vectorizer.transform(texts)
    return matrix.toarray().tolist()


_ensure_qdrant_collection()


def _load_api_key():
    # 1) 优先读系统环境变量
    for key_name in ("DASHSCOPE_API_KEY", "QWEN_API_KEY"):
        value = os.getenv(key_name, "").strip()
        if value:
            return value

    # 2) 再尝试读取 ai_service/.env（支持 uv 启动时未自动注入环境变量）
    env_path = os.path.join(BASE_DIR, ".env")
    if os.path.exists(env_path):
        try:
            with open(env_path, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    key, val = line.split("=", 1)
                    key = key.strip()
                    val = val.strip().strip('"').strip("'")
                    if key in ("DASHSCOPE_API_KEY", "QWEN_API_KEY") and val:
                        return val
        except Exception:
            return ""
    return ""


# 通义千问 API 密钥（未配置时自动降级为本地模式）
dashscope.api_key = _load_api_key()


class RAGService:
    SCORE_THRESHOLD = 0.54
    MIN_FALLBACK_SCORE = 0.16
    FALLBACK_TOP_K = 12
    MAX_STRONG_CHUNKS = 30
    MAX_CONTEXT_CHARS = 7600
    MAX_PREVIEW_ROWS_PER_SHEET = 80
    FIELD_PATTERNS = {
        "Exam_Score": [r"Exam_Score", r"Exam Score", r"考试成绩", r"成绩"],
        "Hours_Studied": [r"Hours_Studied", r"Hours Studied", r"学习时长", r"学习时间"],
        "Sleep_Hours": [r"Sleep_Hours", r"Sleep Hours", r"睡眠时长", r"睡眠时间"],
        "Previous_Scores": [r"Previous_Scores", r"Previous Scores", r"之前成绩", r"历史成绩"],
        "Motivation_Level": [r"Motivation_Level", r"Motivation Level", r"学习动力", r"动力水平"],
    }

    @staticmethod
    def _serialize_dataset(dataset: RagDataset) -> dict[str, Any]:
        return {
            "id": dataset.id,
            "file_name": dataset.file_name,
            "file_path": dataset.file_path,
            "upload_time": dataset.upload_time.strftime("%Y-%m-%d %H:%M:%S"),
            "file_size": dataset.file_size,
        }

    @staticmethod
    def _build_point_id(dataset_id, row_index):
        return int(dataset_id) * 1_000_000 + int(row_index)

    @staticmethod
    def _split_text_content(raw_text, chunk_size=500, overlap=80):
        content = (raw_text or "").strip()
        if not content:
            return []
        chunks = []
        buffer = []
        current_len = 0
        for line in content.splitlines():
            line = line.strip()
            if not line:
                continue
            if current_len + len(line) + 1 > chunk_size and buffer:
                chunk = "\n".join(buffer).strip()
                if chunk:
                    chunks.append(chunk)
                # 保留尾部重叠内容，增强跨段检索稳定性
                tail = chunk[-overlap:] if overlap > 0 else ""
                buffer = [tail, line] if tail else [line]
                current_len = len("\n".join(buffer))
            else:
                buffer.append(line)
                current_len += len(line) + 1
        if buffer:
            chunk = "\n".join(buffer).strip()
            if chunk:
                chunks.append(chunk)
        return chunks

    @staticmethod
    def _normalize_excel_value(value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        text = str(value).strip()
        return "" if text.lower() in {"nan", "none", "null"} else text

    @staticmethod
    def _build_excel_row_keywords(row_map: dict[str, str]) -> list[str]:
        keywords = []
        key_columns = {
            "studentNo",
            "studentName",
            "classCode",
            "courseName",
            "chapterCode",
            "chapterName",
            "knowledgePoint",
            "questionId",
        }
        for key, value in row_map.items():
            if key in key_columns and value:
                keywords.append(value)
        return list(dict.fromkeys(keywords))

    @staticmethod
    def _extract_excel_entries(file_path):
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        entries = []
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows = worksheet.iter_rows(values_only=True)
                try:
                    header_row = next(rows)
                except StopIteration:
                    continue
                headers = [RAGService._normalize_excel_value(item) or f"column_{idx + 1}" for idx, item in enumerate(header_row)]
                for row_index, row in enumerate(rows, start=2):
                    row_map = {}
                    for idx, cell in enumerate(row):
                        header = headers[idx] if idx < len(headers) else f"column_{idx + 1}"
                        value = RAGService._normalize_excel_value(cell)
                        if value:
                            row_map[header] = value
                    if not row_map:
                        continue
                    field_parts = [f"{key}：{value}" for key, value in row_map.items()]
                    entries.append(
                        {
                            "text": f"工作表：{sheet_name}；行号：{row_index}；" + "；".join(field_parts),
                            "sheet_name": sheet_name,
                            "row_index": row_index,
                            "keywords": RAGService._build_excel_row_keywords(row_map),
                        }
                    )
        finally:
            workbook.close()
        return entries

    @staticmethod
    def _extract_texts_from_file(file_path):
        ext = os.path.splitext(file_path)[1].lower()
        if ext in SUPPORTED_EXCEL_EXT:
            return RAGService._extract_excel_entries(file_path), "excel"
        if ext == SUPPORTED_TXT_EXT:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                raw_text = f.read()
            chunks = RAGService._split_text_content(raw_text)
            return [
                {"text": item, "sheet_name": "", "row_index": idx + 1, "keywords": []}
                for idx, item in enumerate(chunks)
            ], "txt"
        raise ValueError("仅支持 .xlsx / .xls / .txt 文件")

    @staticmethod
    def process_file(file_path):
        """解析文件并存入Qdrant和MySQL数据库（支持Excel/TXT）"""
        entries, file_type = RAGService._extract_texts_from_file(file_path)
        if not entries:
            return "文件解析后没有可用文本，请检查内容是否为空"
        file_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path) / 1024  # KB

        db = SessionLocal()
        try:
            dataset = RagDataset(
                file_name=file_name,
                file_path=file_path,
                upload_time=datetime.now(),
                file_size=file_size,
                is_deleted=ACTIVE_FLAG,
            )
            db.add(dataset)
            db.commit()
            db.refresh(dataset)
            dataset_id = dataset.id
        except Exception as e:
            db.rollback()
            return f"数据库记录失败：{str(e)}"
        finally:
            db.close()

        try:
            texts = [item.get("text", "") for item in entries]
            vectors = _embed_texts(texts)
            points = []
            for idx, (entry, vector) in enumerate(zip(entries, vectors)):
                points.append(
                    qmodels.PointStruct(
                        id=RAGService._build_point_id(dataset_id, idx),
                        vector=vector,
                        payload={
                            "dataset_id": dataset_id,
                            "row_index": entry.get("row_index", idx),
                            "file_name": file_name,
                            "sheet_name": entry.get("sheet_name", ""),
                            "keywords": entry.get("keywords", []),
                            "text": entry.get("text", ""),
                        },
                    )
                )
            if points:
                qdrant_client.upsert(collection_name=QDRANT_COLLECTION, points=points, wait=True)
            return f"成功导入 {len(entries)} 条{file_type}数据，数据集ID：{dataset_id}"
        except Exception as e:
            db2 = SessionLocal()
            try:
                rollback_dataset = db2.query(RagDataset).filter(RagDataset.id == dataset_id).first()
                if rollback_dataset:
                    rollback_dataset.is_deleted = DELETED_FLAG
                    db2.commit()
            finally:
                db2.close()
            return f"向量写入失败（数据集ID：{dataset_id}）：{str(e)}"

    @staticmethod
    def process_excel(file_path):
        """兼容旧接口，内部统一走 process_file"""
        return RAGService.process_file(file_path)

    @staticmethod
    def reindex_dataset(dataset_id: int):
        db = None
        try:
            db = SessionLocal()
            dataset = db.query(RagDataset).filter(
                RagDataset.id == dataset_id,
                RagDataset.is_deleted == ACTIVE_FLAG,
            ).first()
            if not dataset:
                return {"error": "数据集不存在或已删除"}

            entries, file_type = RAGService._extract_texts_from_file(dataset.file_path)
            if not entries:
                return {"error": "数据集文件解析后没有可用文本"}

            texts = [item.get("text", "") for item in entries]
            vectors = _embed_texts(texts)
            points = []
            for idx, (entry, vector) in enumerate(zip(entries, vectors)):
                points.append(
                    qmodels.PointStruct(
                        id=RAGService._build_point_id(dataset.id, idx),
                        vector=vector,
                        payload={
                            "dataset_id": dataset.id,
                            "row_index": entry.get("row_index", idx),
                            "file_name": dataset.file_name,
                            "sheet_name": entry.get("sheet_name", ""),
                            "keywords": entry.get("keywords", []),
                            "text": entry.get("text", ""),
                        },
                    )
                )
            for start in range(0, len(points), 2000):
                qdrant_client.upsert(
                    collection_name=QDRANT_COLLECTION,
                    points=points[start:start + 2000],
                    wait=True,
                )
            return {
                "message": f"成功重建 {len(entries)} 条{file_type}索引",
                "datasetId": dataset.id,
                "count": len(entries),
                "fileName": dataset.file_name,
            }
        except Exception as e:
            return {"error": str(e)}
        finally:
            if db is not None:
                db.close()

    @staticmethod
    def _local_fallback_answer(context, business_context=""):
        context_text = (context or "").strip()
        business_text = (business_context or "").strip()
        if business_text:
            lines = []
            if "最新成绩预测：暂无" in business_text or "请先提醒学生去成绩预测模块完成最新预测" in business_text:
                lines.append("当前还没有最新成绩预测结果，建议先去成绩预测模块完成一次预测，再结合预测结果细化提分方案。")
            lines.append("我已先结合当前学生画像做综合分析：")
            for raw_line in business_text.splitlines():
                line = raw_line.strip()
                if not line:
                    continue
                if line.startswith("回答要求") or line[:2] in {"1.", "2.", "3."}:
                    continue
                if any(tag in line for tag in ["刷题总览", "最薄弱章节", "薄弱知识点", "高频错题线索", "编程题表现", "考试成绩"]):
                    lines.append(line)
            if len(lines) > 1:
                return "\n".join(lines[:7])
        if not context_text:
            return "未检索到相关知识。"

        lines = [line.strip() for line in context_text.splitlines() if line.strip()]
        sections = {"概念定义": "", "典型考法": "", "易错点": "", "解题步骤": ""}
        for line in lines:
            for key in sections.keys():
                if key in line and not sections[key]:
                    if "：" in line:
                        sections[key] = line.split("：", 1)[1].strip()
                    elif ":" in line:
                        sections[key] = line.split(":", 1)[1].strip()
                    else:
                        sections[key] = line.strip()

        ordered_blocks = [f"{key}：{value}" for key, value in sections.items() if value]
        if ordered_blocks:
            return "\n\n".join(ordered_blocks[:4])

        advice = []
        for line in lines:
            for tag in ("优化方案：", "调整方案：", "提升方案：", "维持方案：", "建议："):
                if tag in line:
                    part = line.split(tag, 1)[1].strip("；;，, ")
                    if part:
                        advice.append(part)
                    break
            if len(advice) >= 3:
                break

        if not advice:
            preview = context_text[:300]
            return f"已命中相关资料。结论：当前信息可用于分析，建议优先围绕薄弱项做针对训练。\n依据摘要：{preview}"

        unique_advice = []
        for item in advice:
            if item not in unique_advice:
                unique_advice.append(item)
        bullets = [f"{idx + 1}. {text}" for idx, text in enumerate(unique_advice[:3])]
        return "结论：已命中可执行的提分建议。\n建议：\n" + "\n".join(bullets)

    @staticmethod
    def _is_analysis_question(question: str) -> bool:
        q = (question or "").strip()
        keywords = [
            "怎么提高",
            "如何提高",
            "怎么提分",
            "如何提分",
            "学情",
            "分析",
            "提升建议",
            "学习建议",
            "学习规划",
            "复盘",
            "改进",
            "补强",
            "薄弱",
            "针对我",
            "结合我的",
            "结合我当前",
            "教学方案",
            "教学计划",
            "授课方案",
            "教学策略",
            "分层教学",
            "个性化教学",
            "班级方案",
            "课堂讲评",
            "学习资源",
            "资源推荐",
            "推荐资源",
            "资料推荐",
            "推荐资料",
        ]
        return any(keyword in q for keyword in keywords) or any(
            keyword in q
            for keyword in [
                "怎么提高",
                "如何提高",
                "怎么提分",
                "如何提分",
                "学情",
                "分析",
                "提升建议",
                "学习建议",
                "学习规划",
                "复盘",
                "改进",
                "补强",
                "薄弱",
                "针对我",
                "结合我的",
                "结合我当前",
                "教学方案",
                "教学计划",
                "授课方案",
                "教学策略",
                "分层教学",
                "个性化教学",
                "班级方案",
                "课堂讲评",
                "学习资源",
                "资源推荐",
                "推荐资源",
                "资料推荐",
                "推荐资料",
                "区别",
                "差距",
                "对比",
                "相比",
                "比起来",
                "掌握得不好",
                "哪些章节",
                "做得不好",
            ]
        )

    @staticmethod
    def _is_knowledge_explanation_question(question: str) -> bool:
        q = (question or "").strip()
        keywords = [
            "什么是",
            "解释",
            "讲解",
            "题目",
            "这道题",
            "选择题",
            "判断题",
            "答案",
            "为什么选",
            "解题",
            "概念",
            "知识点",
            "考法",
            "易错点",
            "步骤",
        ]
        return any(keyword in q for keyword in keywords)

    @staticmethod
    def _is_teacher_business_context(business_context: str) -> bool:
        text = (business_context or "").strip()
        if not text:
            return False
        return any(
            marker in text
            for marker in [
                "班级整体概览",
                "班级薄弱章节",
                "重点关注学生/模块",
                "班级高频错题",
                "班级薄弱知识点",
                "作业表现概览",
                "教师画像",
            ]
        )

    @staticmethod
    def _is_teacher_class_plan_question(question: str, business_context: str = "") -> bool:
        q = (question or "").strip()
        if not q or not RAGService._is_teacher_business_context(business_context):
            return False
        class_scope = any(word in q for word in ["班级", "全班", "学生整体", "整体学情", "数据班", "这个班"])
        plan_intent = any(
            word in q
            for word in [
                "教学方案",
                "教学计划",
                "授课方案",
                "教学策略",
                "教学安排",
                "课堂安排",
                "讲评方案",
                "分层教学",
                "个性化教学",
                "怎么教",
                "如何教",
            ]
        )
        return class_scope and plan_intent

    @staticmethod
    def _important_teacher_context_lines(business_context: str) -> list[str]:
        lines = []
        for raw_line in (business_context or "").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("回答要求"):
                continue
            if any(
                marker in line
                for marker in [
                    "当前优先分析范围",
                    "班级整体概览",
                    "班级薄弱章节",
                    "重点关注学生/模块",
                    "AI 互动概览",
                    "班级高频错题",
                    "班级薄弱知识点",
                    "作业表现概览",
                ]
            ):
                lines.append(line)
        return lines

    @staticmethod
    def _build_teacher_class_plan_prompt(question: str, context: str, business_context: str = ""):
        return [
            {
                "role": "system",
                "content": (
                    "你是教师端班级学情分析与教学设计助手。当前任务是为教师生成班级层面的个性化教学方案，"
                    "不是解释单道题，也不是罗列某个学生的一条作答记录。"
                    "必须优先使用【教师班级画像】中的班级整体概览、薄弱章节、重点关注学生/模块、高频错题、薄弱知识点和作业表现。"
                    "【知识库命中资料】只能作为补充证据；如果命中的是单个学生的原始作答记录，不要把它当成全班结论，"
                    "只能说它是个别样例，最终判断仍以班级画像为准。"
                    "回答要直接给教师可执行方案，禁止输出 studentNo、answerRecordId、questionId 这类原始字段堆砌。"
                    "固定结构输出："
                    "班级总体判断：用2-3句话说明这个班当前最需要解决的问题。"
                    "数据依据：引用3条以上班级画像中的具体依据。"
                    "分层教学方案：把学生分成基础巩固、重点提升、拓展强化三类，分别写教学动作。"
                    "课堂实施安排：给出课前、课中、课后的安排。"
                    "训练与作业设计：给出针对薄弱章节/知识点的练习策略。"
                    "跟踪反馈：说明下一轮如何用正确率、错题变化、预测成绩或AI互动记录评估效果。"
                    "语言要像教师备课方案，具体、可落地、不要空泛。"
                ),
            },
            {
                "role": "user",
                "content": (
                    f"【教师班级画像】\n{business_context or '暂无教师班级画像'}\n\n"
                    f"【知识库命中资料】\n{context or '暂无命中资料'}\n\n"
                    f"【教师问题】\n{question}"
                ),
            },
        ]

    @staticmethod
    def _build_teacher_class_plan_fallback(question: str, context: str, business_context: str = "") -> str:
        teacher_lines = RAGService._important_teacher_context_lines(business_context)
        if not teacher_lines:
            preview = (context or "").strip()[:260]
            return (
                "班级总体判断：当前没有拿到完整班级画像，只能先根据已命中的知识库资料做保守建议。"
                "建议先导入教师端前序数据，包含班级正确率、薄弱章节、高频错题和学生分层后，再生成完整教学方案。\n\n"
                f"可参考资料：{preview or '暂无可用资料'}"
            )

        evidence = "\n".join([f"{idx + 1}. {line}" for idx, line in enumerate(teacher_lines[:6])])
        weak_focus = next((line for line in teacher_lines if "班级薄弱章节" in line or "班级薄弱知识点" in line), "")
        student_focus = next((line for line in teacher_lines if "重点关注学生/模块" in line), "")
        hot_wrong = next((line for line in teacher_lines if "班级高频错题" in line), "")

        focus_text = weak_focus or hot_wrong or student_focus or teacher_lines[0]
        return (
            "班级总体判断：这次问题应按班级整体学情来处理，而不是依据单个学生的一条作答记录下结论。"
            f"从教师画像看，教学重点应优先围绕“{focus_text}”展开，先解决共性薄弱点，再对不同学生做分层补强。\n\n"
            f"数据依据：\n{evidence}\n\n"
            "分层教学方案：基础巩固层重点补齐概念和基础题，建议用短讲解加即时练习确认是否真正听懂；"
            "重点提升层围绕薄弱章节和高频错题做变式训练，要求学生写出错因和改法；"
            "拓展强化层可以安排综合题或讲题任务，让掌握较好的学生承担示范和互助角色。\n\n"
            "课堂实施安排：课前用系统统计结果确定本节课的薄弱章节和高频错题；课中先讲共性错误，再用分层练习区分基础巩固、重点提升和拓展强化任务；"
            "课后根据新一轮作答正确率和错题变化更新学生分层，继续调整下一次训练内容。\n\n"
            "训练与作业设计：作业不建议一刀切。对薄弱章节布置基础题和同类变式题，对反复出错的知识点增加错题复盘，"
            "对掌握较好的学生布置综合应用题，保证不同水平学生都有对应提升路径。\n\n"
            "跟踪反馈：下一轮重点看班级平均正确率、薄弱章节正确率、高频错题重复出现次数，以及学生AI问答和练习记录是否减少同类错误。"
        )

    @staticmethod
    def _is_resource_recommendation_question(question: str) -> bool:
        q = (question or "").strip()
        if not q:
            return False
        resource_words = ["学习资源", "资源推荐", "推荐资源", "资料推荐", "推荐资料", "课程资源", "练习资源", "视频", "网课", "书籍", "题库"]
        intent_words = ["推荐", "给我", "找", "适合", "基于", "根据", "能否", "可以"]
        return any(word in q for word in resource_words) and any(word in q for word in intent_words)

    @staticmethod
    def _is_learning_schedule_question(question: str) -> bool:
        q = (question or "").strip()
        if not q:
            return False
        time_words = [
            "两周",
            "2周",
            "二周",
            "14天",
            "十四天",
            "一周",
            "7天",
            "七天",
            "每天",
            "每日",
            "本周",
            "下周",
        ]
        schedule_words = [
            "学习安排",
            "学习计划",
            "学习规划",
            "训练安排",
            "复习安排",
            "复习计划",
            "怎么安排",
            "先做什么",
            "优先做什么",
            "任务",
        ]
        return any(word in q for word in time_words) and any(word in q for word in schedule_words)

    @staticmethod
    def _important_student_context_lines(business_context: str) -> list[str]:
        lines = []
        for raw_line in (business_context or "").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("回答要求"):
                continue
            if any(
                marker in line
                for marker in [
                    "最新成绩预测",
                    "刷题总览",
                    "最薄弱章节",
                    "薄弱知识点",
                    "高频错题线索",
                    "编程题表现",
                    "考试成绩",
                    "当前优先分析范围",
                ]
            ):
                lines.append(line)
        return lines

    @staticmethod
    def _build_learning_resource_prompt(question: str, context: str, business_context: str = ""):
        return [
            {
                "role": "system",
                "content": (
                    "你是个性化学习资源推荐助手。当前任务是根据学生真实答题情况、薄弱章节、错题线索和编程表现推荐学习资源。"
                    "知识库命中内容如果是课程资料、教材说明、知识点讲义或资源清单，可以作为依据；"
                    "如果命中的是 studentName、studentNo、answerRecordId、questionId、submitTime、answerContent 等单条作答明细，不能使用里面的姓名称呼用户，不能把它当作当前用户身份。"
                    "除非学生画像里明确给出当前学生姓名，否则统一称呼“同学”，不要编造或沿用知识库片段里的姓名。"
                    "如果知识库没有直接给出资源名称，也不能停在“没有资料”，必须基于学生画像和通用学习经验给出可访问的学习链接、搜索关键词、练习方向和使用方法。"
                    "不要把原始字段直接堆给用户，不要只复述画像。"
                    "推荐外部资源时只能使用下面列出的白名单入口链接，不能编造更深层的课程URL、题单URL或教程URL；"
                    "如果需要更具体，只写“在该站内搜索：关键词”，不要生成未知链接。"
                    "白名单入口：中国大学MOOC https://www.icourse163.org/；学堂在线 https://www.xuetangx.com/；"
                    "华为人才在线 https://www.huawei.com/cn/learning/；菜鸟教程 https://www.runoob.com/；"
                    "力扣题库 https://leetcode.cn/problemset/；牛客题库 https://www.nowcoder.com/exam/oj；B站搜索 https://search.bilibili.com/。"
                    "不要说“我发现你不会”，只能说“当前数据显示某章节/知识点需要优先巩固”。"
                    "不要输出 Markdown 加粗符号、列表缩进符号或 HTML，直接用自然中文分段和编号。"
                    "回答结构固定为："
                    "推荐结论：先说明当前最该补哪类资源。"
                    "依据：引用学生画像中至少3条具体依据。"
                    "资源清单：给出4-6类资源，每类必须包含资源名称、链接或搜索关键词、适合解决什么问题、建议怎么使用。"
                    "练习安排：给出最近3天的资源使用顺序和练习量。"
                    "反馈标准：说明如何判断这些资源是否有效。"
                    "语言要直接、具体、像老师给学生配学习资料。"
                ),
            },
            {
                "role": "user",
                "content": (
                    f"【学生画像/业务数据】\n{business_context or '暂无学生画像'}\n\n"
                    f"【知识库命中资料】\n{context or '暂无直接命中的资源资料'}\n\n"
                    f"【学生问题】\n{question}"
                ),
            },
        ]

    @staticmethod
    def _filter_learning_resource_context(context: str) -> str:
        text = (context or "").strip()
        if not text:
            return ""
        raw_lines = [line.strip() for line in text.splitlines() if line.strip()]
        blocked_markers = [
            "studentName",
            "studentNo",
            "answerRecordId",
            "questionId",
            "submitTime",
            "answerContent",
            "standardAnswer",
            "isCorrect",
        ]
        useful_lines = []
        for line in raw_lines:
            if any(marker in line for marker in blocked_markers):
                continue
            useful_lines.append(line)
        filtered = "\n".join(useful_lines).strip()
        if not filtered and any(marker in text for marker in blocked_markers):
            return ""
        return filtered[:3600]

    @staticmethod
    def _build_learning_resource_fallback(question: str, context: str, business_context: str = "") -> str:
        student_lines = RAGService._important_student_context_lines(business_context)
        evidence = "\n".join([f"{idx + 1}. {line}" for idx, line in enumerate(student_lines[:6])])
        weak_line = next((line for line in student_lines if "最薄弱章节" in line or "薄弱知识点" in line), "")
        wrong_line = next((line for line in student_lines if "高频错题线索" in line), "")
        program_line = next((line for line in student_lines if "编程题表现" in line), "")
        focus = weak_line or wrong_line or program_line or "当前答题记录中暴露出的薄弱章节和高频错题"
        if not evidence:
            evidence = "1. 当前未拿到完整学生画像，建议先导入历史作答、诊断结果和预测成绩后再细化资源。"

        return (
            f"推荐结论：建议优先围绕“{focus}”配置资源，不要泛泛刷题。知识库没有直接给出可用资源名称时，可以先使用通用资源类型和关键词检索来补齐。\n\n"
            f"依据：\n{evidence}\n\n"
            "资源清单：\n"
            "1. 中国大学MOOC：https://www.icourse163.org/，搜索“计算机基础”“数据结构”“人工智能基础”，用于系统补基础概念。\n"
            "2. 学堂在线：https://www.xuetangx.com/，搜索“数据结构”“程序设计基础”，优先选择高校课程做章节复习。\n"
            "3. 华为人才在线：https://www.huawei.com/cn/learning/，搜索“ICT”“人工智能”“HCIA”，用于匹配华为ICT-AI相关知识。\n"
            "4. 菜鸟教程：https://www.runoob.com/，搜索 C / Python / SQL / 数据结构相关教程，用来快速查漏补缺。\n"
            "5. 力扣题库：https://leetcode.cn/problemset/，按“链表、数组、哈希表、栈队列”等标签做编程实践。\n"
            "6. 牛客题库：https://www.nowcoder.com/exam/oj，用来做计算机基础、编程题和选择题训练。\n"
            "7. B站搜索：https://search.bilibili.com/，搜索“章节名 + 易错题/选择题解析/数据结构动画”，只看有例题过程的讲解。\n\n"
            "练习安排：第1天先看基础概念和2-3道例题；第2天做同类变式题10道并整理错因；第3天复做错题并补一组综合题。"
            "如果连续两天同类题正确率能达到80%以上，再进入下一章节或更高难度资源。\n\n"
            "反馈标准：看薄弱章节正确率是否提升、同类错题是否减少、是否能独立说出错因和解题步骤。"
        )

    @staticmethod
    def _build_learning_schedule_prompt(question: str, context: str, business_context: str = ""):
        return [
            {
                "role": "system",
                "content": (
                    "你是学生个性化学习规划助手。当前任务是根据学生画像和命中资料，直接生成学习周期安排。"
                    "用户问两周、一周、每天、优先做什么时，绝对不能只复述画像摘要，必须给出可执行日程。"
                    "必须优先使用学生画像中的成绩预测、刷题总量、正确率、最薄弱章节、薄弱知识点、高频错题线索、编程题表现和考试成绩。"
                    "如果缺少考试或最新预测，可以说明缺失，但仍然要基于已有作答数据继续制定计划。"
                    "如果知识库命中的是原始作答明细，只作为错题线索，不要把字段原文堆给用户。"
                    "回答结构固定为："
                    "总体目标：用1-2句话说明这段周期要解决什么。"
                    "数据依据：列出至少3条来自学生画像的具体依据。"
                    "两周安排：按第1天到第14天列出，每天都要写优先任务、训练量、复盘动作。"
                    "每天优先级：说明每天先做什么、后做什么。"
                    "检查标准：说明第7天和第14天分别看哪些指标。"
                    "不要输出 Markdown 标题符号、不要输出 ###、##、**、HTML 标签。"
                    "语言要像老师给学生布置任务，具体、清楚、可照做。"
                ),
            },
            {
                "role": "user",
                "content": (
                    f"【学生画像/业务数据】\n{business_context or '暂无学生画像'}\n\n"
                    f"【补充资料】\n{context or '暂无命中资料'}\n\n"
                    f"【学生问题】\n{question}"
                ),
            },
        ]

    @staticmethod
    def _build_learning_schedule_fallback(question: str, context: str, business_context: str = "") -> str:
        student_lines = RAGService._important_student_context_lines(business_context)
        evidence = "\n".join([f"{idx + 1}. {line}" for idx, line in enumerate(student_lines[:6])])
        weak_line = next((line for line in student_lines if "最薄弱章节" in line or "薄弱知识点" in line), "")
        wrong_line = next((line for line in student_lines if "高频错题线索" in line), "")
        prediction_line = next((line for line in student_lines if "最新成绩预测" in line), "")
        practice_line = next((line for line in student_lines if "刷题总览" in line), "")
        focus = weak_line or wrong_line or "当前答题记录中的薄弱章节与高频错题"
        if not evidence:
            evidence = "1. 当前没有拿到完整学生画像，建议先导入历史作答、诊断结果和成绩预测后再细化。"

        return (
            "总体目标：接下来两周先不盲目刷题，重点围绕当前最薄弱章节和高频错题做“基础回补 + 同类变式 + 错题复盘”。"
            f"优先解决“{focus}”，让同类题正确率稳定提升。\n\n"
            f"数据依据：\n{evidence}\n\n"
            "两周安排：\n"
            "第1天：梳理薄弱章节，把错题按知识点分组；完成基础题10道，记录每道错因。\n"
            "第2天：复习薄弱章节核心概念；做同类选择题15道，错题必须写出正确答案依据。\n"
            "第3天：围绕高频错题线索做变式训练10道；复做第1天错题，检查是否还会错。\n"
            "第4天：补当前薄弱知识点的例题讲解；做基础题10道、提升题5道。\n"
            "第5天：做一次小测15-20题；把错误集中到“概念不清、审题失误、步骤遗漏”三类。\n"
            "第6天：针对小测错因专项补强；每类错因至少完成5道同类题。\n"
            "第7天：周复盘，只做错题重做和薄弱章节回测；目标是同类题正确率达到75%以上。\n"
            "第8天：进入第二轮巩固，先复习第1周仍错的知识点；做变式题15道。\n"
            "第9天：加入综合题训练，把薄弱知识点放到综合场景中练；完成10道综合题。\n"
            "第10天：如果有编程/实践题，安排1个小任务；重点写清思路、步骤和调试记录。\n"
            "第11天：回到选择题和判断题，限时完成20题；训练速度和稳定性。\n"
            "第12天：整理两周错题本，只保留仍不稳定的题型；每个题型写一句避错提醒。\n"
            "第13天：做一次综合模拟，控制时间；结束后只分析错题，不再新增大量内容。\n"
            "第14天：最终复盘，重做第7天和第13天错题；总结下阶段还要补的2个知识点。\n\n"
            "每天优先级：先看错题和薄弱知识点，再做同类基础题，最后做少量提升题。"
            "每天最后10分钟必须写复盘：今天错在哪里、明天先补什么、哪类题已经稳定。\n\n"
            "检查标准：第7天看薄弱章节正确率是否接近或超过75%，同类错题是否明显减少；"
            "第14天看综合题正确率、错题复错率和答题速度是否改善。"
            f"{' 参考当前预测：' + prediction_line if prediction_line else ''}"
            f"{' 参考刷题情况：' + practice_line if practice_line else ''}"
        )

    @staticmethod
    def _normalize_header_name(header: str) -> str:
        text = (header or "").strip()
        if not text:
            return ""
        return re.sub(r"[\s_\-:/（）()]+", "", text).lower()

    @staticmethod
    def _find_row_value(row_map: dict[str, str], aliases: list[str]) -> str:
        if not row_map:
            return ""
        normalized_map = {
            RAGService._normalize_header_name(key): str(value).strip()
            for key, value in row_map.items()
            if str(value).strip()
        }
        for alias in aliases:
            value = normalized_map.get(RAGService._normalize_header_name(alias), "")
            if value:
                return value
        return ""

    @staticmethod
    def _parse_float(value: Any) -> float | None:
        text = str(value or "").strip().rstrip("%")
        if not text:
            return None
        try:
            return float(text)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _parse_int(value: Any) -> int | None:
        number = RAGService._parse_float(value)
        if number is None:
            return None
        return int(round(number))

    @staticmethod
    def _load_active_datasets(active_ids: list[int]) -> list[dict[str, Any]]:
        if not active_ids:
            return []
        db = None
        try:
            db = SessionLocal()
            datasets = db.query(RagDataset).filter(
                RagDataset.id.in_(active_ids),
                RagDataset.is_deleted == ACTIVE_FLAG,
            ).all()
            return [
                {
                    "id": item.id,
                    "file_name": item.file_name,
                    "file_path": item.file_path,
                }
                for item in datasets
            ]
        finally:
            if db is not None:
                db.close()

    @staticmethod
    def _is_student_summary_sheet(sheet_name: str) -> bool:
        name = (sheet_name or "").strip().lower()
        return (
            "学生章节汇总" in name
            or ("学生" in name and "章节" in name and "汇总" in name)
            or ("student" in name and "chapter" in name and "summary" in name)
        )

    @staticmethod
    def _extract_student_names_for_analysis(question: str, active_ids: list[int]) -> list[str]:
        q = (question or "").strip()
        if not q:
            return []
        matched_names = []
        for dataset in RAGService._load_active_datasets(active_ids):
            file_path = dataset.get("file_path", "")
            ext = os.path.splitext(file_path)[1].lower()
            if ext not in SUPPORTED_EXCEL_EXT or not os.path.exists(file_path):
                continue
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            try:
                for sheet_name in workbook.sheetnames:
                    if not RAGService._is_student_summary_sheet(sheet_name):
                        continue
                    rows = workbook[sheet_name].iter_rows(values_only=True)
                    try:
                        headers = next(rows)
                    except StopIteration:
                        continue
                    header_names = [
                        RAGService._normalize_excel_value(item) or f"column_{idx + 1}"
                        for idx, item in enumerate(headers)
                    ]
                    for row in rows:
                        row_map = {}
                        for idx, cell in enumerate(row):
                            key = header_names[idx] if idx < len(header_names) else f"column_{idx + 1}"
                            value = RAGService._normalize_excel_value(cell)
                            if value:
                                row_map[key] = value
                        student_name = RAGService._find_row_value(
                            row_map,
                            ["studentName", "student_name", "学生姓名", "姓名", "学生"],
                        )
                        if student_name and student_name in q and student_name not in matched_names:
                            matched_names.append(student_name)
                        if len(matched_names) >= 3:
                            return matched_names
            finally:
                workbook.close()
        return matched_names

    @staticmethod
    def _load_student_summary_records(active_ids: list[int], target_names: list[str]) -> list[dict[str, Any]]:
        if not active_ids or not target_names:
            return []
        target_set = set(target_names)
        records: list[dict[str, Any]] = []
        for dataset in RAGService._load_active_datasets(active_ids):
            file_path = dataset.get("file_path", "")
            ext = os.path.splitext(file_path)[1].lower()
            if ext not in SUPPORTED_EXCEL_EXT or not os.path.exists(file_path):
                continue
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            try:
                for sheet_name in workbook.sheetnames:
                    if not RAGService._is_student_summary_sheet(sheet_name):
                        continue
                    rows = workbook[sheet_name].iter_rows(values_only=True)
                    try:
                        headers = next(rows)
                    except StopIteration:
                        continue
                    header_names = [
                        RAGService._normalize_excel_value(item) or f"column_{idx + 1}"
                        for idx, item in enumerate(headers)
                    ]
                    for row_index, row in enumerate(rows, start=2):
                        row_map = {}
                        for idx, cell in enumerate(row):
                            key = header_names[idx] if idx < len(header_names) else f"column_{idx + 1}"
                            value = RAGService._normalize_excel_value(cell)
                            if value:
                                row_map[key] = value
                        if not row_map:
                            continue
                        student_name = RAGService._find_row_value(
                            row_map,
                            ["studentName", "student_name", "学生姓名", "姓名", "学生"],
                        )
                        if student_name not in target_set:
                            continue
                        chapter_name = RAGService._find_row_value(
                            row_map,
                            ["chapterName", "chapter_name", "章节名称", "章节"],
                        )
                        if not chapter_name:
                            continue
                        answer_count = RAGService._parse_int(
                            RAGService._find_row_value(row_map, ["answerCount", "answer_count", "作答数", "答题数", "题量"])
                        )
                        correct_count = RAGService._parse_int(
                            RAGService._find_row_value(row_map, ["correctCount", "correct_count", "正确数", "答对数"])
                        )
                        wrong_count = RAGService._parse_int(
                            RAGService._find_row_value(row_map, ["wrongCount", "wrong_count", "错误数", "答错数"])
                        )
                        correct_rate = RAGService._parse_float(
                            RAGService._find_row_value(row_map, ["correctRate", "correct_rate", "正确率", "得分率"])
                        )
                        if correct_rate is None and answer_count and correct_count is not None:
                            correct_rate = round(correct_count * 100.0 / answer_count, 2)
                        if correct_rate is None:
                            continue
                        records.append(
                            {
                                "datasetId": dataset["id"],
                                "fileName": dataset["file_name"],
                                "sheetName": sheet_name,
                                "rowIndex": row_index,
                                "studentName": student_name,
                                "studentNo": RAGService._find_row_value(
                                    row_map, ["studentNo", "student_no", "学号", "学生学号"]
                                ),
                                "courseName": RAGService._find_row_value(
                                    row_map, ["courseName", "course_name", "课程名称", "课程"]
                                ) or "未标注课程",
                                "chapterName": chapter_name,
                                "answerCount": answer_count or 0,
                                "correctCount": correct_count if correct_count is not None else 0,
                                "wrongCount": wrong_count if wrong_count is not None else max((answer_count or 0) - (correct_count or 0), 0),
                                "correctRate": round(correct_rate, 2),
                            }
                        )
            finally:
                workbook.close()
        return records

    @staticmethod
    def _build_teacher_analysis_sources(records: list[dict[str, Any]], limit: int = 8) -> list[dict[str, Any]]:
        ordered = sorted(records, key=lambda item: (item["studentName"], item["correctRate"], -item["answerCount"]))
        sources = []
        for item in ordered[:limit]:
            snippet = (
                f"学生：{item['studentName']}；课程：{item['courseName']}；章节：{item['chapterName']}；"
                f"正确率：{item['correctRate']}%；作答数：{item['answerCount']}；错误数：{item['wrongCount']}"
            )
            sources.append(
                {
                    "datasetId": item["datasetId"],
                    "fileName": item["fileName"],
                    "sheetName": item["sheetName"],
                    "rowIndex": item["rowIndex"],
                    "score": 1.0,
                    "snippet": snippet,
                }
            )
        return sources

    @staticmethod
    def _build_teacher_student_analysis_answer(question: str, records: list[dict[str, Any]]) -> str:
        if not records:
            return ""
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for item in records:
            grouped[item["studentName"]].append(item)
        student_names = sorted(
            grouped.keys(),
            key=lambda name: ((question or "").find(name) if (question or "").find(name) >= 0 else 999999, name),
        )
        compare_mode = len(student_names) >= 2 and any(
            keyword in (question or "")
            for keyword in ["对比", "相比", "差距", "区别", "比起来", "更差", "更弱"]
        )

        lines = []
        if compare_mode:
            lines.append(f"当前判断：你这次问的是学生之间的差距和区别，不是知识点讲解。我直接按章节汇总数据比较 {student_names[0]} 和 {student_names[1]}。")
        else:
            lines.append(f"当前判断：你这次问的是学生学情，不是题目解析。我直接按章节汇总数据分析 {student_names[0]} 的薄弱章节。")

        for student_name in student_names[:2]:
            candidates = [item for item in grouped[student_name] if item["answerCount"] >= 3]
            if not candidates:
                candidates = grouped[student_name]
            weakest = sorted(candidates, key=lambda x: (x["correctRate"], -x["answerCount"], x["chapterName"]))[:3]
            lines.append(f"{student_name}薄弱章节：")
            for row in weakest:
                lines.append(
                    f"{row['courseName']} 的“{row['chapterName']}”正确率 {row['correctRate']}%，作答 {row['answerCount']} 题，错 {row['wrongCount']} 题。"
                )

        if compare_mode:
            left_name, right_name = student_names[:2]
            right_index = {
                (item["courseName"], item["chapterName"]): item
                for item in grouped[right_name]
            }
            gaps = []
            for item in grouped[left_name]:
                pair = right_index.get((item["courseName"], item["chapterName"]))
                if not pair:
                    continue
                gap = round(item["correctRate"] - pair["correctRate"], 2)
                gaps.append((abs(gap), gap, item, pair))
            gaps.sort(key=lambda x: (-x[0], -max(x[2]["answerCount"], x[3]["answerCount"])))
            if gaps:
                top_gaps = gaps[:3]
                lines.append(f"{left_name}与{right_name}的主要差距：")
                for _, gap, left_item, right_item in top_gaps:
                    if gap < 0:
                        lines.append(
                            f"在 {left_item['courseName']} 的“{left_item['chapterName']}”上，{left_name} 正确率 {left_item['correctRate']}%，低于 {right_name} 的 {right_item['correctRate']}%，差距 {abs(gap)} 个百分点。"
                        )
                    elif gap > 0:
                        lines.append(
                            f"在 {left_item['courseName']} 的“{left_item['chapterName']}”上，{left_name} 正确率 {left_item['correctRate']}%，高于 {right_name} 的 {right_item['correctRate']}%，领先 {abs(gap)} 个百分点。"
                        )
                weaker_gaps = sorted([item for item in gaps if item[1] < 0], key=lambda x: -x[0])
                if weaker_gaps:
                    focus = weaker_gaps[0][2]
                    if top_gaps and all(item[1] >= 0 for item in top_gaps):
                        lines.append(
                            f"教学建议：整体上 {left_name} 并不落后，但在 {focus['courseName']} 的“{focus['chapterName']}”仍低于 {right_name}，这块最值得单点补强。"
                        )
                    else:
                        lines.append(
                            f"教学建议：优先补 {left_name} 在 {focus['courseName']} 的“{focus['chapterName']}”，这里是当前最明显的量化短板。"
                        )
                else:
                    lines.append(f"教学建议：{left_name} 并不是整体落后，更适合做局部章节补强。")
            else:
                lines.append("对比说明：当前命中的汇总数据里，两位同学没有足够的同课程同章节交集，所以只能分别判断薄弱点，不能硬编差距。")
        else:
            candidates = [item for item in grouped[student_names[0]] if item["answerCount"] >= 3]
            if not candidates:
                candidates = grouped[student_names[0]]
            weakest = sorted(candidates, key=lambda x: (x["correctRate"], -x["answerCount"]))[:2]
            if weakest:
                lines.append(
                    f"教学建议：先补“{weakest[0]['chapterName']}”，再补“{weakest[min(1, len(weakest) - 1)]['chapterName']}”，因为这两处同时满足正确率偏低且作答量不小，不像偶然失误。"
                )

        return "\n".join(lines)

    @staticmethod
    def _try_answer_student_analysis(question: str, active_ids: list[int]) -> dict[str, Any] | None:
        if not RAGService._is_analysis_question(question):
            return None
        if RAGService._is_knowledge_explanation_question(question):
            return None
        student_names = RAGService._extract_student_names_for_analysis(question, active_ids)
        if not student_names:
            return None
        records = RAGService._load_student_summary_records(active_ids, student_names)
        if not records:
            return None
        answer = RAGService._build_teacher_student_analysis_answer(question, records)
        if not answer:
            return None
        sources = RAGService._build_teacher_analysis_sources(records)
        return {
            "answer": answer,
            "mode": "student_analysis",
            "sources": sources,
            "matchedCount": len(sources),
            "matchedDatasetCount": len({src.get("datasetId") for src in sources if src.get("datasetId") is not None}),
        }

    @staticmethod
    def _local_general_answer(question):
        q = (question or "").strip()
        if any(x in q for x in ["你好", "您好", "hi", "hello"]):
            return "你好，我是教育平台 AI 助手。你可以问我学习规划、作业建议、考试复盘和成绩提升。"
        if "作业" in q:
            return "作业建议：先做基础题，再做提升题；每道错题记录“错因+改法”，第二天复做一次。"
        if "考试" in q or "复习" in q:
            return "考试复习建议：先梳理考点清单，再按薄弱点分配时间，最后做1-2套限时模拟并复盘错题。"
        if "成绩" in q or "提高" in q:
            return "提分建议：固定每天学习时段，先补薄弱科目；每周做一次错题复盘并跟踪正确率变化。"
        return "当前未连接在线大模型。你可以继续提问，我会按教育场景给你结构化建议。"

    @staticmethod
    def _call_qwen(messages, timeout=15):
        def call_model():
            return call_deepseek_chat(messages, timeout=timeout)

        with ThreadPoolExecutor(max_workers=1) as executor:
            return executor.submit(call_model).result(timeout=timeout)

    @staticmethod
    def _chat_without_knowledge(question):
        if not is_deepseek_available():
            return RAGService._local_general_answer(question)
        try:
            answer = RAGService._call_qwen(
                [
                    {
                        "role": "system",
                        "content": (
                            "你是教育场景的AI助手，请直接、清晰、自然地回答用户问题。"
                            "如果用户问的是选择题、判断题或题目解析，优先按“正确答案 + 解析”作答。"
                            "解析要稍微展开一些，至少说明为什么该答案正确，并简要说明其他明显干扰项为什么不对。"
                            "不要只给一两句过短答案，也不要写得像提示词说明。"
                        ),
                    },
                    {"role": "user", "content": question},
                ]
            )
            return answer or RAGService._local_general_answer(question)
        except TimeoutError:
            return RAGService._local_general_answer(question)
        except Exception as e:
            return f"{RAGService._local_general_answer(question)}\n（在线模型调用异常：{str(e)}）"

    @staticmethod
    def query_answer(question, active_ids=None):
        result = RAGService.query_answer_with_meta(question, active_ids=active_ids)
        return result.get("answer", "")

    @staticmethod
    def _build_source_item(payload: dict[str, Any], score: float):
        text = str(payload.get("text", "")).strip()
        return {
            "datasetId": payload.get("dataset_id"),
            "fileName": payload.get("file_name", ""),
            "sheetName": payload.get("sheet_name", ""),
            "rowIndex": payload.get("row_index"),
            "score": round(score, 4),
            "snippet": text[:180],
            "text": text,
        }

    @staticmethod
    def _pick_diverse_sources(sources: list[dict[str, Any]], limit: int) -> list[dict[str, Any]]:
        if not sources or limit <= 0:
            return []
        return sorted(sources, key=lambda x: x.get("score", 0), reverse=True)[:limit]

    @staticmethod
    def _extract_query_terms(question: str) -> list[str]:
        raw = (question or "").strip()
        if not raw:
            return []
        stop_terms = {
            "哪些", "哪个", "怎么", "如何", "情况", "同学", "学生", "老师", "班级", "课程", "章节", "掌握", "不好",
            "较差", "差距", "起来", "比较", "表现", "分析", "一下", "以及", "然后", "现在", "这个", "那个",
        }
        terms = []
        for token in re.findall(r"\d{5,}|[A-Za-z][A-Za-z0-9_-]{1,20}", raw):
            token = token.strip()
            if token and token not in stop_terms:
                terms.append(token)
        for segment in re.findall(r"[\u4e00-\u9fff]{2,}", raw):
            segment = segment.strip()
            if not segment:
                continue
            max_len = min(6, len(segment))
            for size in range(2, max_len + 1):
                for start in range(0, len(segment) - size + 1):
                    token = segment[start:start + size]
                    if token in stop_terms:
                        continue
                    if token.endswith(("哪些", "哪里", "情况", "比较", "分析")):
                        continue
                    terms.append(token)
        return list(dict.fromkeys(terms))

    @staticmethod
    def _score_keyword_bonus(text: str, keywords: list[str]) -> float:
        if not text or not keywords:
            return 0.0
        bonus = 0.0
        for keyword in keywords:
            if keyword and keyword in text:
                if re.fullmatch(r"\d{5,}", keyword):
                    bonus += 0.18
                elif re.fullmatch(r"[\u4e00-\u9fff]{2,6}", keyword):
                    bonus += 0.14
                else:
                    bonus += 0.08
        return min(bonus, 0.42)

    @staticmethod
    def _retrieve_context(question: str, active_ids: list[int]) -> dict[str, Any]:
        query_terms = RAGService._extract_query_terms(question)
        candidate_results = []
        query_texts = [question] + query_terms[:6]
        query_limits = [180] + [80] * min(len(query_terms), 6)
        for query_text, limit in zip(query_texts, query_limits):
            if not query_text:
                continue
            query_vec = _embed_texts([query_text])[0]
            result = qdrant_client.query_points(
                collection_name=QDRANT_COLLECTION,
                query=query_vec,
                limit=limit,
                with_payload=True,
            )
            candidate_results.extend(getattr(result, "points", []) or [])

        active_set = set(active_ids)
        dedup = {}
        strong_sources = []
        fallback_sources = []
        for item in candidate_results:
            payload = item.payload or {}
            dataset_id = payload.get("dataset_id")
            if dataset_id not in active_set:
                continue
            point_key = (dataset_id, payload.get("sheet_name", ""), payload.get("row_index"))
            raw_score = float(getattr(item, "score", 0.0) or 0.0)
            final_score = raw_score + RAGService._score_keyword_bonus(str(payload.get("text", "")), query_terms)
            previous = dedup.get(point_key)
            if previous and previous.get("score", 0) >= final_score:
                continue
            source = RAGService._build_source_item(payload, final_score)
            dedup[point_key] = source
        for source in dedup.values():
            if not source["text"]:
                continue
            score = float(source.get("score", 0.0) or 0.0)
            if score >= RAGService.SCORE_THRESHOLD:
                strong_sources.append(source)
            if score >= RAGService.MIN_FALLBACK_SCORE:
                fallback_sources.append(source)

        if strong_sources:
            selected = RAGService._pick_diverse_sources(strong_sources, RAGService.MAX_STRONG_CHUNKS)
            mode = "strong"
        else:
            selected = RAGService._pick_diverse_sources(fallback_sources, RAGService.FALLBACK_TOP_K)
            mode = "fallback" if selected else "none"

        texts = []
        total_chars = 0
        for src in selected:
            seg = src["text"]
            if total_chars + len(seg) > RAGService.MAX_CONTEXT_CHARS:
                break
            texts.append(seg)
            total_chars += len(seg)

        context = "\n".join(texts)
        return {
            "mode": mode,
            "context": context,
            "sources": selected,
        }

    @staticmethod
    def _build_rag_prompt(question: str, context: str):
        return [
            {
                "role": "system",
                "content": (
                    "你是教育平台RAG助手。仅依据给定资料回答，不要编造，不要沿用与当前问题无关的旧上下文。"
                    "如果检索片段与当前问题不直接相关，必须明确说明资料不匹配，而不是勉强作答。"
                    "当资料中已经出现可用于分析的学生字段（如 Exam_Score、Hours_Studied、Sleep_Hours、Previous_Scores 等）时，"
                    "必须先给出可执行建议，不要直接回答“无法给建议”。"
                    "输出必须是总结性表达，先给结论，再给 3-5 条编号建议，避免复述检索片段原文。"
                    "只有在完全没有相关字段时，才说明信息不足，并明确需要补充哪些字段。"
                ),
            },
            {"role": "user", "content": f"【资料】\n{context}\n\n【问题】\n{question}"},
        ]

    @staticmethod
    def _build_natural_rag_prompt(question: str, context: str, business_context: str = ""):
        return [
            {
                "role": "system",
                "content": (
                    "你是教育场景 RAG 助手。"
                    "只能依据当前检索到的资料回答，不要混入无关历史上下文，不要编造知识点。"
                    "如果命中了课程、模块或知识点资料，回答必须优先围绕资料中的模块说明组织。"
                    "如果资料中出现“该知识点属于……模块”“RAG命中后应优先围绕……”这类教学指引，必须严格按该指引的回答顺序输出。"
                    "默认回答顺序为：概念定义、典型考法、易错点、解题步骤。"
                    "如果题目本身是选择题或判断题，可以在上述四部分之后再补充答案结论。"
                    "如果同时给了学生业务画像，且用户在问学情分析、提升建议、学习规划、复盘建议，必须明确引用画像里的成绩预测、刷题、错题、编程题、考试数据。"
                    "如果业务画像里明确写了暂无最新成绩预测，第一句先提醒先去做成绩预测，再结合现有数据继续分析。"
                    "如果知识库与学生画像同时存在，先按知识库规则解释知识点，再衔接个性化建议。"
                    "不要输出 Markdown 标题符号，不要输出 ###、##、**、HTML 标签。"
                    "直接输出自然中文分段内容，每一部分都要写出对应标题文本，例如“概念定义：”“典型考法：”“易错点：”“解题步骤：”。"
                    "如果资料与问题不完全匹配，要明确说明匹配不足，但仍然优先解释当前最相关的知识点。"
                ),
            },
            {
                "role": "user",
                "content": (
                    f"【学生业务画像】\n{business_context or '暂无业务画像'}\n\n"
                    f"【知识资料】\n{context or '暂无命中知识资料'}\n\n"
                    f"【问题】\n{question}"
                ),
            },
        ]

    @staticmethod
    def _build_analysis_prompt(question: str, context: str, business_context: str = ""):
        return [
            {
                "role": "system",
                "content": (
                    "你是学生学情深度分析助手。"
                    "当前任务不是讲知识点定义，而是结合学生画像认真回答“我现在问题在哪、为什么会这样、先补什么、具体怎么做”。"
                    "必须优先使用学生业务画像中的成绩预测、刷题情况、薄弱章节、薄弱知识点、错题、高频失分点、编程题表现、考试成绩。"
                    "如果业务画像中写明“最新成绩预测：暂无”，第一句先提醒学生先去做一次成绩预测，但后面仍然要继续结合现有数据分析，不能只提醒一句就结束。"
                    "如果同时命中了课程知识资料，只把它当成补充依据，用来解释为什么某个章节或知识点要优先补，不要把回答主体写成概念讲解。"
                    "回答必须有足够深度，不要只写四五行。"
                    "整段回答总字数尽量保持在 450 到 700 字之间，每个部分至少写 2 句话。"
                    "回答结构固定为："
                    "当前判断：用 2-3 句话概括学生当前最主要的状态与问题。"
                    "核心依据：至少引用 3 条具体数据，明确写出正确率、作答量、最近练习情况、均分或错题线索。"
                    "问题拆解：分别说明基础薄弱、练习频率、错题处理、知识点掌握这几个方面哪里出了问题。"
                    "提升重点：给出 3-4 个按优先级排序的补强方向，并解释每一项为什么现在最重要。每一点都要写出对应数据依据。"
                    "三天安排：按第1天、第2天、第3天给出可执行训练计划，每天要写训练内容、建议题量、复盘重点。"
                    "风险提醒：补一句如果继续保持当前状态，最可能卡分的地方是什么。"
                    "如果缺少考试表现或最新预测，要明确说明当前场景里暂时没有这项数据，但不能因为缺少一项数据就停止分析。"
                    "不要只重复画像原文，要把数据翻译成原因判断和行动建议。"
                    "如果资料里出现具体知识点或章节，只在“问题拆解”和“提升重点”里点出它为什么要补、怎么补，不要长篇复述概念。"
                    "不要输出 Markdown 标题符号，不要输出 ##、###、** 或 HTML 标签。"
                    "语言要像老师看完数据后给出的详细诊断，直接、具体、别说空话。"
                ),
            },
            {
                "role": "user",
                "content": (
                    f"【学生业务画像】\n{business_context or '暂无业务画像'}\n\n"
                    f"【补充知识资料】\n{context or '暂无命中知识资料'}\n\n"
                    f"【用户问题】\n{question}"
                ),
            },
        ]

    @staticmethod
    def _needs_analysis_expansion(answer: str) -> bool:
        text = (answer or "").strip()
        if not text:
            return True
        required_sections = ["当前判断：", "核心依据：", "问题拆解：", "提升重点：", "三天安排：", "风险提醒："]
        if any(section not in text for section in required_sections):
            return True
        return len(text) < 520

    @staticmethod
    def _rewrite_analysis_answer(question: str, context: str, business_context: str, draft_answer: str) -> str:
        prompt = [
            {
                "role": "system",
                "content": (
                    "你是学生学情分析润色助手。"
                    "你的任务不是重写成空话，而是把已有分析补全到更深入、更具体、更像老师真实诊断。"
                    "请在已有草稿基础上补充，不要缩短。"
                    "必须保留并输出以下标题：当前判断、核心依据、问题拆解、提升重点、三天安排、风险提醒。"
                    "每个部分至少 2 句话，总字数不少于 450 字。"
                    "核心依据至少引用 3 条具体数据。"
                    "提升重点必须写出优先级和原因。"
                    "三天安排必须分别写训练内容、建议题量、复盘重点。"
                    "不要输出 Markdown 标题符号，不要输出 ##、###、** 或 HTML 标签。"
                    "如果没有最新成绩预测或考试表现，要点明缺失项，再继续结合已有数据分析。"
                ),
            },
            {
                "role": "user",
                "content": (
                    f"【学生业务画像】\n{business_context or '暂无业务画像'}\n\n"
                    f"【补充知识资料】\n{context or '暂无命中知识资料'}\n\n"
                    f"【用户问题】\n{question}\n\n"
                    f"【已有草稿】\n{draft_answer or '暂无草稿'}"
                ),
            },
        ]
        return RAGService._call_qwen(prompt)

    @staticmethod
    def _contains_insufficient_signal(answer: str) -> bool:
        text = (answer or "").strip()
        if not text:
            return True
        patterns = [
            "无法直接得出",
            "无法给出",
            "无法提供",
            "信息不足",
            "没有提供",
            "缺少",
        ]
        return any(p in text for p in patterns)

    @staticmethod
    def _extract_numeric_field(all_text: str, aliases: list[str]) -> int | None:
        for alias in aliases:
            escaped = re.escape(alias)
            regex_list = [
                rf"{escaped}\s*[：:=]\s*[\"']?(-?\d+(?:\.\d+)?)",
                rf"[\"']{escaped}[\"']\s*:\s*[\"']?(-?\d+(?:\.\d+)?)",
            ]
            for reg in regex_list:
                match = re.search(reg, all_text, flags=re.IGNORECASE)
                if match:
                    try:
                        return int(float(match.group(1)))
                    except (TypeError, ValueError):
                        continue
        return None

    @staticmethod
    def _extract_student_fields_from_text(all_text: str) -> dict[str, int]:
        fields: dict[str, int] = {}
        for key, aliases in RAGService.FIELD_PATTERNS.items():
            value = RAGService._extract_numeric_field(all_text, aliases)
            if value is not None:
                fields[key] = value
        return fields

    @staticmethod
    def _rule_summary_from_sources(question: str, sources: list[dict[str, Any]]) -> str:
        all_text = "\n".join([str((src or {}).get("text", "")).strip() for src in sources if src])
        if not all_text:
            return ""

        fields = RAGService._extract_student_fields_from_text(all_text)
        score = fields.get("Exam_Score")
        hours = fields.get("Hours_Studied")
        sleep = fields.get("Sleep_Hours")
        prev = fields.get("Previous_Scores")
        motivation = fields.get("Motivation_Level")

        advice = []
        if hours is not None and hours > 10:
            advice.append("把长时低效学习压缩到 6-8 小时，采用“2小时学习+10分钟复盘”的节奏，优先做错题与薄弱点。")
        if sleep is not None:
            if sleep > 9:
                advice.append("将睡眠逐步调整到 7-8 小时，把新增时间用于薄弱学科专项练习。")
            elif 6 <= sleep <= 8:
                advice.append("保持 6-8 小时规律睡眠，固定作息，保障注意力稳定。")
            else:
                advice.append("先把睡眠补到 6-8 小时，避免疲劳导致学习效率下滑。")
        if prev is not None and score is not None and score <= prev + 10:
            advice.append("按周跟踪同科目分项得分，定位掉分题型后做针对训练，避免只刷总量。")
        if motivation is not None and motivation <= 4:
            advice.append("把学习任务切成 25-30 分钟小目标并设置即时反馈，先恢复学习启动意愿。")
        if score is not None and score < 75:
            advice.append("优先保证基础题满分率，先清理高频错题再做综合题，避免一上来刷难题。")

        if not advice:
            lines = [line.strip() for line in all_text.splitlines() if line.strip()]
            for line in lines:
                for tag in ("优化方案：", "调整方案：", "提升方案：", "维持方案：", "建议："):
                    if tag in line:
                        part = line.split(tag, 1)[1].strip("；;，, ")
                        if part and part not in advice:
                            advice.append(part)
                        break
                if len(advice) >= 3:
                    break

        if not advice:
            return ""

        conclusion = "结论：有可落地的提分路径，核心是提升学习效率与策略匹配。"
        if score is not None:
            conclusion = f"结论：当前成绩约为 {score} 分，具备明确提升空间。"
        items = "\n".join([f"{idx + 1}. {item}" for idx, item in enumerate(advice[:4])])
        return f"{conclusion}\n建议：\n{items}"

    @staticmethod
    def query_answer_with_meta(question, active_ids=None, business_context=""):
        if active_ids is None:
            db = None
            try:
                db = SessionLocal()
                active_ids = [item.id for item in db.query(RagDataset.id).filter(RagDataset.is_deleted == ACTIVE_FLAG).all()]
            finally:
                if db is not None:
                    db.close()

        business_text = (business_context or "").strip()
        teacher_class_plan = RAGService._is_teacher_class_plan_question(question, business_text)
        resource_recommendation = RAGService._is_resource_recommendation_question(question)
        learning_schedule = RAGService._is_learning_schedule_question(question)

        if not active_ids:
            direct_answer = RAGService._chat_without_knowledge(question)
            if learning_schedule and business_text and is_deepseek_available():
                try:
                    direct_answer = RAGService._call_qwen(
                        RAGService._build_learning_schedule_prompt(question, "", business_text),
                        timeout=25,
                    )
                except Exception:
                    direct_answer = RAGService._build_learning_schedule_fallback(question, "", business_text)
            elif resource_recommendation and business_text and is_deepseek_available():
                try:
                    direct_answer = RAGService._call_qwen(
                        RAGService._build_learning_resource_prompt(question, "", business_text),
                        timeout=25,
                    )
                except Exception:
                    direct_answer = RAGService._build_learning_resource_fallback(question, "", business_text)
            return {
                "answer": (
                    RAGService._build_teacher_class_plan_fallback(question, "", business_text)
                    if teacher_class_plan else
                    RAGService._build_learning_schedule_fallback(question, "", business_text)
                    if learning_schedule and business_text and not is_deepseek_available() else
                    RAGService._build_learning_resource_fallback(question, "", business_text)
                    if resource_recommendation and business_text and not is_deepseek_available() else
                    direct_answer
                    if (learning_schedule or resource_recommendation) and business_text else
                    (
                        "\u5f53\u524d\u8fd8\u6ca1\u6709\u68c0\u7d22\u5230\u53ef\u7528\u7684\u77e5\u8bc6\u5e93\u8d44\u6599\uff0c\u4f46\u5df2\u63a5\u5165\u5b66\u60c5\u753b\u50cf\uff0c\u5148\u6309\u5b66\u751f\u6570\u636e\u4e3a\u4f60\u5206\u6790\uff1a\n\n"
                        + RAGService._local_fallback_answer("", business_text)
                    ) if business_text else (
                        "\u5f53\u524d\u8fd8\u6ca1\u6709\u68c0\u7d22\u5230\u53ef\u7528\u7684\u77e5\u8bc6\u5e93\u8d44\u6599\uff0c\u5148\u4e3a\u4f60\u76f4\u63a5\u7ed9\u51fa\u7b54\u6848\uff1a\n\n"
                        + direct_answer
                    )
                ),
                "mode": "chat",
                "sources": [],
                "matchedCount": 0,
            }

        analysis_direct = None if learning_schedule else RAGService._try_answer_student_analysis(question, active_ids)
        if analysis_direct:
            return analysis_direct

        try:
            retrieved = RAGService._retrieve_context(question, active_ids)
        except Exception as e:
            return {
                "answer": f"\u77e5\u8bc6\u5e93\u68c0\u7d22\u5931\u8d25\uff1a{str(e)}",
                "mode": "error",
                "sources": [],
                "matchedCount": 0,
            }

        context = retrieved.get("context", "")
        sources = retrieved.get("sources", [])
        mode = retrieved.get("mode", "none")
        analysis_question = RAGService._is_analysis_question(question)
        resource_context = RAGService._filter_learning_resource_context(context) if resource_recommendation else context
        if resource_recommendation and not resource_context:
            sources = []
            mode = "resource_recommendation"
        if not context and not business_text:
            direct_answer = RAGService._chat_without_knowledge(question)
            return {
                "answer": (
                    "\u672a\u68c0\u7d22\u5230\u76f4\u63a5\u76f8\u5173\u7684\u77e5\u8bc6\u5e93\u5185\u5bb9\uff0c\u4e0b\u9762\u5148\u6309\u901a\u7528\u80fd\u529b\u4e3a\u4f60\u89e3\u7b54\uff1a\n\n"
                    f"{direct_answer}"
                ),
                "mode": mode,
                "sources": [],
                "matchedCount": 0,
            }

        merged_context = context

        if not is_deepseek_available():
            unique_dataset_count = len({src.get("datasetId") for src in sources if src.get("datasetId") is not None})
            return {
                "answer": (
                    RAGService._build_teacher_class_plan_fallback(question, merged_context, business_text)
                    if teacher_class_plan
                    else RAGService._build_learning_schedule_fallback(question, merged_context, business_text)
                    if learning_schedule and business_text
                    else RAGService._build_learning_resource_fallback(question, resource_context, business_text)
                    if resource_recommendation and business_text
                    else RAGService._local_fallback_answer(merged_context, business_text)
                ),
                "mode": mode,
                "sources": [{k: v for k, v in src.items() if k != "text"} for src in sources],
                "matchedCount": len(sources),
                "matchedDatasetCount": unique_dataset_count,
            }

        try:
            prompt = (
                RAGService._build_teacher_class_plan_prompt(question, merged_context, business_text)
                if teacher_class_plan
                else
                RAGService._build_learning_schedule_prompt(question, merged_context, business_text)
                if learning_schedule and business_text
                else
                RAGService._build_learning_resource_prompt(question, resource_context, business_text)
                if resource_recommendation and business_text
                else
                RAGService._build_analysis_prompt(question, merged_context, business_text)
                if analysis_question and business_text
                else RAGService._build_natural_rag_prompt(question, merged_context, business_text)
            )
            answer = RAGService._call_qwen(prompt, timeout=25 if (resource_recommendation or learning_schedule) else 15)
            if (not teacher_class_plan) and (not learning_schedule) and (not resource_recommendation) and analysis_question and business_text and RAGService._needs_analysis_expansion(answer):
                expanded_answer = RAGService._rewrite_analysis_answer(question, merged_context, business_text, answer)
                if expanded_answer:
                    answer = expanded_answer
        except TimeoutError:
            answer = (
                RAGService._build_teacher_class_plan_fallback(question, merged_context, business_text)
                if teacher_class_plan
                else RAGService._build_learning_schedule_fallback(question, merged_context, business_text)
                if learning_schedule and business_text
                else RAGService._build_learning_resource_fallback(question, resource_context, business_text)
                if resource_recommendation and business_text
                else RAGService._local_fallback_answer(merged_context, business_text)
            )
        except Exception as e:
            fallback_answer = (
                RAGService._build_teacher_class_plan_fallback(question, merged_context, business_text)
                if teacher_class_plan
                else RAGService._build_learning_schedule_fallback(question, merged_context, business_text)
                if learning_schedule and business_text
                else RAGService._build_learning_resource_fallback(question, resource_context, business_text)
                if resource_recommendation and business_text
                else RAGService._local_fallback_answer(merged_context, business_text)
            )
            answer = f"{fallback_answer}\n在线模型调用异常：{str(e)}"

        if (not learning_schedule) and (not resource_recommendation) and RAGService._contains_insufficient_signal(answer):
            fallback_summary = RAGService._rule_summary_from_sources(question, sources)
            if fallback_summary:
                answer = fallback_summary

        unique_dataset_count = len({src.get("datasetId") for src in sources if src.get("datasetId") is not None})
        return {
            "answer": answer or (
                RAGService._build_teacher_class_plan_fallback(question, merged_context, business_text)
                if teacher_class_plan
                else RAGService._build_learning_schedule_fallback(question, merged_context, business_text)
                if learning_schedule and business_text
                else RAGService._build_learning_resource_fallback(question, resource_context, business_text)
                if resource_recommendation and business_text
                else RAGService._local_fallback_answer(merged_context, business_text)
            ),
            "mode": mode,
            "sources": [{k: v for k, v in src.items() if k != "text"} for src in sources],
            "matchedCount": len(sources),
            "matchedDatasetCount": unique_dataset_count,
        }

    @staticmethod
    def get_datasets():
        """获取所有未删除的数据集"""
        db = None
        try:
            db = SessionLocal()
            datasets = db.query(RagDataset).filter(RagDataset.is_deleted == ACTIVE_FLAG).all()
            return [RAGService._serialize_dataset(dataset) for dataset in datasets]
        except Exception as e:
            return {"error": str(e)}
        finally:
            if db is not None:
                db.close()

    @staticmethod
    def delete_dataset(dataset_id):
        """删除数据集（软删除）"""
        db = None
        try:
            db = SessionLocal()
            dataset = db.query(RagDataset).filter(RagDataset.id == dataset_id).first()
            if not dataset:
                return {"error": "数据集不存在"}

            dataset.is_deleted = DELETED_FLAG
            db.commit()
            return {"message": "数据集删除成功"}
        except Exception as e:
            if db is not None:
                db.rollback()
            return {"error": str(e)}
        finally:
            if db is not None:
                db.close()

    @staticmethod
    def get_dataset_detail(dataset_id):
        """获取数据集详情"""
        db = None
        try:
            db = SessionLocal()
            dataset = db.query(RagDataset).filter(
                RagDataset.id == dataset_id,
                RagDataset.is_deleted == ACTIVE_FLAG,
            ).first()

            if not dataset:
                return {"error": "数据集不存在或已删除"}

            ext = os.path.splitext(dataset.file_path)[1].lower()
            if ext in SUPPORTED_EXCEL_EXT:
                data = []
                workbook = load_workbook(dataset.file_path, read_only=True, data_only=True)
                try:
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        rows = worksheet.iter_rows(values_only=True)
                        try:
                            header_row = next(rows)
                        except StopIteration:
                            continue
                        headers = [RAGService._normalize_excel_value(item) or f"column_{idx + 1}" for idx, item in enumerate(header_row)]
                        sheet_rows = []
                        for idx, row in enumerate(rows, start=1):
                            if idx > RAGService.MAX_PREVIEW_ROWS_PER_SHEET:
                                break
                            row_map = {}
                            for col_idx, cell in enumerate(row):
                                header = headers[col_idx] if col_idx < len(headers) else f"column_{col_idx + 1}"
                                value = RAGService._normalize_excel_value(cell)
                                if value:
                                    row_map[header] = value
                            if row_map:
                                sheet_rows.append(row_map)
                        if sheet_rows:
                            data.append({"sheetName": sheet_name, "rows": sheet_rows})
                finally:
                    workbook.close()
            elif ext == SUPPORTED_TXT_EXT:
                with open(dataset.file_path, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read().strip()
                chunks = RAGService._split_text_content(content, chunk_size=300)
                data = [{"content": item} for item in chunks[:50]]
            else:
                data = [{"content": "暂不支持该文件类型的详情预览"}]

            result = RAGService._serialize_dataset(dataset)
            result["data"] = data
            return result
        except Exception as e:
            return {"error": str(e)}
        finally:
            if db is not None:
                db.close()
